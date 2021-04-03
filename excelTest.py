import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import question
import getpass
import os


def create_file(testName):

    wb = openpyxl.Workbook()
    filepath = "/Desktop/" + testName + ".xlsx"
    wb.save('C:/Users/' + getpass.getuser() + filepath)

    wb1 = openpyxl.load_workbook('C:/Users/' + getpass.getuser() + filepath)

    mySheet = wb1['Sheet']

    mySheet['A1'] = 'Course Name'
    mySheet['B1'] = 'Test Name'
    mySheet['C1'] = 'Question Text'
    mySheet['D1'] = 'Right Answer 1'
    mySheet['E1'] = 'Right Answer 2'
    mySheet['F1'] = 'Right Answer 3'
    mySheet['G1'] = 'Right Answer 4'
    mySheet['H1'] = 'Right Answer 5'
    mySheet['I1'] = 'Right Answer 6'
    mySheet['J1'] = 'Right Answer 7'
    mySheet['K1'] = 'Right Answer 8'
    mySheet['L1'] = 'Wrong Answer 1'
    mySheet['M1'] = 'Wrong Answer 2'
    mySheet['N1'] = 'Wrong Answer 3'
    mySheet['O1'] = 'Wrong Answer 4'
    mySheet['P1'] = 'Wrong Answer 5'
    mySheet['Q1'] = 'Wrong Answer 6'
    mySheet['R1'] = 'Wrong Answer 7'
    mySheet['S1'] = 'Wrong Answer 8'
    mySheet['T1'] = 'Correct Feedback'
    mySheet['U1'] = 'Incorrect Feedback'

    mySheet.column_dimensions['A'].width = 18
    mySheet.column_dimensions['B'].width = 16
    mySheet.column_dimensions['C'].width = 19
    mySheet.column_dimensions['D'].width = 19
    mySheet.column_dimensions['E'].width = 19
    mySheet.column_dimensions['F'].width = 19
    mySheet.column_dimensions['G'].width = 19
    mySheet.column_dimensions['H'].width = 19
    mySheet.column_dimensions['I'].width = 19
    mySheet.column_dimensions['J'].width = 19
    mySheet.column_dimensions['K'].width = 19
    mySheet.column_dimensions['L'].width = 20
    mySheet.column_dimensions['M'].width = 20
    mySheet.column_dimensions['N'].width = 20
    mySheet.column_dimensions['O'].width = 20
    mySheet.column_dimensions['P'].width = 20
    mySheet.column_dimensions['Q'].width = 20
    mySheet.column_dimensions['R'].width = 20
    mySheet.column_dimensions['S'].width = 20
    mySheet.column_dimensions['T'].width = 22
    mySheet.column_dimensions['U'].width = 24

    thin = Side(border_style="thick", color="303030")
    letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U"]
    for x in letters:
        column = x + "1"
        mySheet[column].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        mySheet[column].font = Font(size=14, bold=True)
        mySheet[column].alignment = Alignment(horizontal="center")
        mySheet[column].fill = PatternFill(start_color="00ffcc", fill_type="solid")
        # Thanks to user "crifan" on Stack Overflow for explaining how to use PatternFill
        # https://stackoverflow.com/questions/35918504/adding-a-background-color-to-cell-openpyxl

    wb1.save('C:/Users/' + getpass.getuser() + filepath)


def load_questions(pool, filepath):

    wb1 = openpyxl.load_workbook(filepath)

    mySheet = wb1['Sheet']
    letters = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U"]

    i = 2
    end = 0
    while True:
        cell = "C" + str(i)
        if str(mySheet[cell].value) == "None":
            end = 1
            break
        else:
            something = question.Question()
            something.text = str(mySheet[cell].value)
            cell = "D" + str(i)
            something.r1 = str(mySheet[cell].value)
            cell = "E" + str(i)
            something.r2 = str(mySheet[cell].value)
            cell = "F" + str(i)
            something.r3 = str(mySheet[cell].value)
            cell = "G" + str(i)
            something.r4 = str(mySheet[cell].value)
            cell = "H" + str(i)
            something.r5 = str(mySheet[cell].value)
            cell = "I" + str(i)
            something.r6 = str(mySheet[cell].value)
            cell = "J" + str(i)
            something.r7 = str(mySheet[cell].value)
            cell = "K" + str(i)
            something.r8 = str(mySheet[cell].value)
            cell = "L" + str(i)
            something.w1 = str(mySheet[cell].value)
            cell = "M" + str(i)
            something.w2 = str(mySheet[cell].value)
            cell = "N" + str(i)
            something.w3 = str(mySheet[cell].value)
            cell = "O" + str(i)
            something.w4 = str(mySheet[cell].value)
            cell = "P" + str(i)
            something.w5 = str(mySheet[cell].value)
            cell = "Q" + str(i)
            something.w6 = str(mySheet[cell].value)
            cell = "R" + str(i)
            something.w7 = str(mySheet[cell].value)
            cell = "S" + str(i)
            something.w8 = str(mySheet[cell].value)
            cell = "T" + str(i)
            something.correct = str(mySheet[cell].value)
            cell = "U" + str(i)
            something.incorrect = str(mySheet[cell].value)
            something.num_right = 0
            if something.r1 != "None":
                something.num_right += 1
            if something.r2 != "None":
                something.num_right += 1
            if something.r3 != "None":
                something.num_right += 1
            if something.r4 != "None":
                something.num_right += 1
            if something.r5 != "None":
                something.num_right += 1
            if something.r6 != "None":
                something.num_right += 1
            if something.r7 != "None":
                something.num_right += 1
            if something.r8 != "None":
                something.num_right += 1

            something.num_wrong = 0
            if something.w1 != "None":
                something.num_wrong += 1
            if something.w2 != "None":
                something.num_wrong += 1
            if something.w3 != "None":
                something.num_wrong += 1
            if something.w4 != "None":
                something.num_wrong += 1
            if something.w5 != "None":
                something.num_wrong += 1
            if something.w6 != "None":
                something.num_wrong += 1
            if something.w7 != "None":
                something.num_wrong += 1
            if something.w8 != "None":
                something.num_wrong += 1

            pool.append(something)

        if end == 1:
            break
        else:
            i = i + 1


def getCourseName(filepath):
    wb1 = openpyxl.load_workbook(filepath)

    mySheet = wb1['Sheet']

    return str(mySheet['A2'].value)


def getTestName(filepath):
    wb1 = openpyxl.load_workbook(filepath)

    mySheet = wb1['Sheet']

    return str(mySheet['B2'].value)


def getRandom():
    option_filepath = find("options.xlsx", "C:/Users")
    wb1 = openpyxl.load_workbook(option_filepath)

    mySheet = wb1['Sheet']

    if str(mySheet['B3'].value) == '1':
        return '1'
    elif str(mySheet['B3'].value) == '0':
        return '0'
    else:
        mySheet['A3'] = 'Question Order Randomization'
        mySheet['B3'] = '1'
        wb1.save(option_filepath)
        return '1'


def getPartialCredit():
    option_filepath = find("options.xlsx", "C:/Users")
    wb1 = openpyxl.load_workbook(option_filepath)

    mySheet = wb1['Sheet']

    if str(mySheet['B4'].value) == '1':
        return '1'
    elif str(mySheet['B4'].value) == '0':
        return '0'
    else:
        mySheet['A4'] = 'Partial Credit for Multiple Answer Questions'
        mySheet['B4'] = '0'
        wb1.save(option_filepath)
        return '0'


def getQuestionOrderStyle():
    option_filepath = find("options.xlsx", "C:/Users")
    wb1 = openpyxl.load_workbook(option_filepath)

    mySheet = wb1['Sheet']

    if str(mySheet['B2'].value) == '1':
        return '1'
    elif str(mySheet['B2'].value) == '2':
        return '2'
    elif str(mySheet['B2'].value) == '3':
        return '3'
    elif str(mySheet['B2'].value) == '4':
        return '4'
    elif str(mySheet['B2'].value) == '5':
        return '5'
    else:
        mySheet['A2'] = 'Answer Numbering'
        mySheet['B2'] = '5'
        wb1.save(option_filepath)
        return '5'


def optionsFile():  # Makes sure options file exists and creates it if it doesn't
    option_filepath = find("options.xlsx", "C:/Users")
    if option_filepath == "None":
        wb = openpyxl.Workbook()
        wb.save('C:/Users/' + getpass.getuser() + '/Downloads/BlackBoardTestMaker-main/options.xlsx')

        wb1 = openpyxl.load_workbook('C:/Users/' + getpass.getuser() +
                                     '/Downloads/BlackBoardTestMaker-main/options.xlsx')

        mySheet = wb1['Sheet']

        mySheet['A1'] = 'Option'
        mySheet['A2'] = 'Answer Numbering'
        mySheet['A3'] = 'Question Order Randomization'
        mySheet['A4'] = 'Partial Credit for Multiple Answer Questions'

        mySheet['B1'] = 'Value'
        mySheet['B2'] = '5'
        mySheet['B3'] = '1'
        mySheet['B4'] = '0'
        wb1.save('C:/Users/' + getpass.getuser() + '/Downloads/BlackBoardTestMaker-main/options.xlsx')
    return option_filepath


def find(name, path):  # This is so the program can search for where the webdriver file is on the user's computer
    placeholder = ""
    for root, dirs, files in os.walk(path):
        if name in files:
            placeholder = os.path.join(root, name)
            break
        else:
            placeholder = "None"

    return placeholder



